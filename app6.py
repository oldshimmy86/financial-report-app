import streamlit as st
import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime, date
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO

# Настройки Streamlit
st.title("Финансовый Отчёт")

st.markdown("""
Программа для сбора финансовых данных и экспорта их в Excel.
Нажмите кнопку ниже, чтобы сгенерировать отчёт.
""")

# Учетные данные (будут использоваться из Streamlit Secrets)
username = st.secrets["username"]
password = st.secrets["password"]

base_url = "https://api.moysklad.ru/api/remap/1.2/entity"

# Выбор периода отчёта
start_date_input = st.date_input("Дата начала", value=date(2022, 1, 1))
end_date_input = st.date_input("Дата окончания", value=date.today())

# Преобразование дат в формат datetime
start_date = datetime.combine(start_date_input, datetime.min.time())
end_date = datetime.combine(end_date_input, datetime.max.time())

# Функция получения заказов
def fetch_orders(order_type):
    orders = []
    limit = 1000
    offset = 0
    while True:
        url = f"{base_url}/{order_type}"
        params = {'limit': limit, 'offset': offset}
        response = requests.get(url, auth=HTTPBasicAuth(username, password), params=params)
        data = response.json()
        orders.extend(order for order in data['rows'] if 'applicable' in order and order['applicable'])
        if len(data['rows']) < limit:
            break
        offset += limit
    return orders

# Обработка данных
def process_data(cashin_data, cashout_data):
    currency_mapping = {
        "currency/e03f64a6-2225-11ed-0a80-073a00365127": "PLN",
        "currency/e15d9c47-2226-11ed-0a80-04b900364797": "USD",
        "currency/e1754d40-cc82-11ec-0a80-08ab00701a1e": "EUR"
    }
    payment_type_mapping = {
        "Card-in-showroom": "card",
        "Cash-in-showroom": "cash"
    }
    currency_totals = {cur: {"cash": 0, "card": 0, "count": 0, "test_count": 0} for cur in currency_mapping.values()}
    details = []

    for data, doc_type in [(cashin_data, "Приход"), (cashout_data, "Расход")]:
        for item in data:
            currency_href = item["rate"]["currency"]["meta"]["href"]
            currency = next((v for k, v in currency_mapping.items() if k in currency_href), "UNKNOWN")
            amount = round(item["sum"] / 100, 2)  # Сумма хранится в копейках
            test_order = False

            # Определение типа платежа через атрибут PaymentType и test_order
            payment_type = None
            for attr in item.get("attributes", []):
                if attr["name"] == "PaymentType":
                    payment_type = payment_type_mapping.get(attr["value"]["name"])
                elif attr["name"] == "test_order":
                    test_order = attr.get("value", False)

            # Пропускаем записи с неизвестным payment_type
            if not payment_type or currency == "UNKNOWN":
                continue

            # Коррекция для расхода
            if doc_type == "Расход":
                amount = -amount

            # Обновляем остатки по валютам
            if test_order:
                currency_totals[currency]["test_count"] += 1
            else:
                currency_totals[currency][payment_type] += amount
                currency_totals[currency]["count"] += 1

            # Добавляем детали
            details.append({
                "date": item["moment"].split(" ")[0],
                "order_number": item["name"],
                "amount": amount,
                "currency": currency,
                "payment_type": payment_type,
                "doc_type": doc_type,
                "test_order": "yes" if test_order else "no",
                "comment": item.get("description", "")
            })
    
    return currency_totals, details

# Создание Excel-отчёта
def create_excel_report(currency_totals, details):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Баланс по валютам"
    ws2 = wb.create_sheet("Детали ордеров")

    # Заполнение первой вкладки
ws1.append(["Валюта", "Наличные", "Карта", "Количество документов", "Тестовые ордера"])
for currency, data in currency_totals.items():
    ws1.append([currency, data["cash"], data["card"], data["count"], data["test_count"]])
    
    # Проверка отрицательных значений и установка стиля
    if data["cash"] < 0 or data["card"] < 0:
        for cell in ws1.iter_rows(min_row=ws1.max_row, max_row=ws1.max_row):
            for c in cell:
                c.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Заполнение второй вкладки
    ws2.append(["Дата", "Номер ордера", "Сумма", "Валюта", "Тип платежа", "Тип документа", "Test Order", "Комментарий"])
    for detail in details:
        ws2.append([
            detail["date"], detail["order_number"], detail["amount"],
            detail["currency"], detail["payment_type"], detail["doc_type"],
            detail["test_order"], detail["comment"]
        ])
    
    # Сохранение в память
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Генерация отчёта
if st.button("Сгенерировать Отчёт"):
    with st.spinner("Генерация отчёта..."):
        cashin_data = fetch_orders("cashin")
        cashout_data = fetch_orders("cashout")
        cashin_data = [order for order in cashin_data if start_date <= datetime.fromisoformat(order["moment"].replace("Z", "")) <= end_date]
        cashout_data = [order for order in cashout_data if start_date <= datetime.fromisoformat(order["moment"].replace("Z", "")) <= end_date]
        currency_totals, details = process_data(cashin_data, cashout_data)
        excel_file = create_excel_report(currency_totals, details)
        st.success("Отчёт успешно сгенерирован!")
        st.download_button(
            label="Скачать Excel-файл",
            data=excel_file,
            file_name="financial_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
