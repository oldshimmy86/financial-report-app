import streamlit as st
import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO

# Настройки Streamlit
st.title("Финансовый Отчёт")

st.markdown("""
Программа для сбора финансовых данных и экспорта их в Excel.
Нажмите кнопку ниже, чтобы сгенерировать отчёт.
""")

# Учетные данные (будут использоваться из Streamlit Secrets)
username = st.secrets["username"]
password = st.secrets["password"]

base_url = "https://api.moysklad.ru/api/remap/1.2/entity"

# Выбор периода отчёта
start_date = st.date_input("Дата начала", value=datetime.strptime("2022-01-01", '%Y-%m-%d'))
end_date = st.date_input("Дата окончания", value=datetime.today())

# Функции
def fetch_orders(order_type):
    orders = []
    limit = 1000
    offset = 0
    while True:
        url = f"{base_url}/{order_type}"
        params = {
            'limit': limit,
            'offset': offset
        }
        response = requests.get(url, auth=HTTPBasicAuth(username, password), params=params)
        data = response.json()
        
        orders.extend(order for order in data['rows'] if order.get('applicable', False))
        
        if len(data['rows']) < limit:
            break
        offset += limit
    return orders

def filter_orders_by_date(orders, start_date, end_date):
    filtered_orders = []
    for order in orders:
        order_date = datetime.strptime(order['moment'], '%Y-%m-%d %H:%M:%S.%f')
        if start_date <= order_date <= end_date:
            filtered_orders.append(order)
    return filtered_orders

def calculate_totals(orders, is_income):
    totals = {'PLN': {'total': 0, 'cash': 0, 'card': 0, 'count': 0},
              'USD': {'total': 0, 'cash': 0, 'card': 0, 'count': 0},
              'EUR': {'total': 0, 'cash': 0, 'card': 0, 'count': 0}}
    order_details = []
    
    for order in orders:
        sum_value = order['sum'] / 100
        sum_value = sum_value if is_income else -sum_value
        currency_href = order['rate']['currency']['meta']['href']
        
        currency = None
        if "currency/e03f64a6-2225-11ed-0a80-073a00365127" in currency_href:
            currency = 'PLN'
        elif "currency/e15d9c47-2226-11ed-0a80-04b900364797" in currency_href:
            currency = 'USD'
        elif "currency/e1754d40-cc82-11ec-0a80-08ab00701a1e" in currency_href:
            currency = 'EUR'
        
        if currency:
            totals[currency]['total'] += sum_value
            totals[currency]['count'] += 1

            payment_type = next((attr['value']['name'] for attr in order.get('attributes', []) if attr['name'] == "PaymentType"), "Unknown")
            if payment_type == "Cash-in-showroom":
                totals[currency]['cash'] += sum_value
            elif payment_type == "Card-in-showroom":
                totals[currency]['card'] += sum_value

            comment = order.get("description", "")

            order_details.append({
                'date': order['moment'].split(' ')[0],
                'name': order['name'],
                'currency': currency,
                'cash_pln': sum_value if payment_type == "Cash-in-showroom" and currency == "PLN" else "",
                'cash_usd': sum_value if payment_type == "Cash-in-showroom" and currency == "USD" else "",
                'cash_eur': sum_value if payment_type == "Cash-in-showroom" and currency == "EUR" else "",
                'card': sum_value if payment_type == "Card-in-showroom" else "",
                'comment': comment
            })
                
    return totals, order_details

def generate_excel():
    cash_in_orders = fetch_orders("cashin")
    cash_out_orders = fetch_orders("cashout")
    
    cash_in_orders = filter_orders_by_date(cash_in_orders, start_date, end_date)
    cash_out_orders = filter_orders_by_date(cash_out_orders, start_date, end_date)
    
    income_totals, income_details = calculate_totals(cash_in_orders, is_income=True)
    expense_totals, expense_details = calculate_totals(cash_out_orders, is_income=False)
    
    all_details = sorted(income_details + expense_details, key=lambda x: x['date'])
    
    wb = Workbook()
    
    summary_sheet = wb.create_sheet(title="Остатки по валютам")
    summary_sheet.append(["Валюта", "Наличные", "Карта", "Количество документов"])
    
    for currency, data in income_totals.items():
        net_total = data['total'] + expense_totals[currency]['total']
        net_cash = data['cash'] + expense_totals[currency]['cash']
        net_card = data['card'] + expense_totals[currency]['card']
        net_count = data['count'] + expense_totals[currency]['count']
        summary_sheet.append([currency, net_cash, net_card, net_count])
    
    details_sheet = wb.create_sheet(title="Детали ордеров")
    details_sheet.append(["Дата", "Номер ордера", "Cash, PLN", "PLN total", "Cash, USD", "Cash, EUR", "Card", "Currency", "Comment"])
    
    details_sheet.column_dimensions['A'].width = 15
    
    pln_total = 0
    
    for detail in all_details:
        cash_pln = detail['cash_pln'] if detail['cash_pln'] else 0
        pln_total += cash_pln

        row = [
            detail['date'], 
            detail['name'], 
            cash_pln if cash_pln != "" else "", 
            pln_total,
            detail['cash_usd'], 
            detail['cash_eur'], 
            detail['card'], 
            detail['currency'],
            detail['comment']
        ]
        details_sheet.append(row)
    
        current_row = details_sheet.max_row
        pln_total_cell = details_sheet.cell(row=current_row, column=4)
        
        if pln_total < 0:
            pln_total_cell.font = Font(color="FF0000")
        else:
            pln_total_cell.font = Font(color="000000")
    
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)
    
    # Сохранение файла в буфер
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

# Кнопка для генерации отчёта
if st.button("Сгенерировать Отчёт"):
    with st.spinner("Генерация отчёта..."):
        excel_file = generate_excel()
        st.success("Отчёт успешно сгенерирован!")
        st.download_button(
            label="Скачать Excel-файл",
            data=excel_file,
            file_name="financial_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
